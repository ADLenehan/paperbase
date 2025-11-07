#!/usr/bin/env python3
"""
Quick validation test for export refactor with complex data types

Run this to validate the backend export functionality before frontend integration.

Usage:
    python test_export_validation.py
"""

import sys
import json
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook

# Mock classes for testing without database
class MockField:
    def __init__(self, name, field_type, value, value_json=None, verified=False, confidence=0.9):
        self.field_name = name
        self.field_type = field_type
        self.field_value = value if field_type in ["text", "number", "date", "boolean"] else None
        self.field_value_json = value_json
        self.verified = verified
        self.verified_value = None
        self.verified_value_json = None
        self.confidence_score = confidence

class MockDocument:
    def __init__(self, doc_id, filename, status="completed"):
        self.id = doc_id
        self.filename = filename
        self.status = status
        self.uploaded_at = None
        self.processed_at = None
        self.extracted_fields = []

    def add_field(self, field):
        self.extracted_fields.append(field)
        return self


def create_test_documents():
    """Create mock documents with various field types"""

    # Document 1: Invoice with array and array_of_objects
    doc1 = MockDocument(1, "invoice_001.pdf", "completed")
    doc1.add_field(MockField("invoice_number", "text", "INV-001"))
    doc1.add_field(MockField("total_amount", "number", "1250.00"))
    doc1.add_field(MockField("tags", "array", None, ["urgent", "paid", "q4"]))
    doc1.add_field(MockField("line_items", "array_of_objects", None, [
        {"description": "Widget A", "quantity": 5, "unit_price": 100.00},
        {"description": "Widget B", "quantity": 10, "unit_price": 75.00}
    ]))

    # Document 2: Garment spec with table
    doc2 = MockDocument(2, "garment_spec_001.pdf", "completed")
    doc2.add_field(MockField("product_name", "text", "T-Shirt Basic"))
    doc2.add_field(MockField("colors", "array", None, ["Red", "Blue", "Green", "Yellow"]))
    doc2.add_field(MockField("grading_table", "table", None, {
        "rows": [
            {"pom_code": "B510", "size_2": 10.5, "size_3": 11.0, "size_4": 11.5},
            {"pom_code": "B520", "size_2": 12.0, "size_3": 12.5, "size_4": 13.0},
            {"pom_code": "B530", "size_2": 14.0, "size_3": 14.5, "size_4": 15.0}
        ]
    }))

    # Document 3: Simple document (no complex fields)
    doc3 = MockDocument(3, "simple_doc.pdf", "completed")
    doc3.add_field(MockField("name", "text", "John Doe"))
    doc3.add_field(MockField("email", "text", "john@example.com"))
    doc3.add_field(MockField("age", "number", "35"))

    return [doc1, doc2, doc3]


def test_detect_complex_fields():
    """Test: Detect complex fields in documents"""
    from app.services.export_service import ExportService

    print("\n" + "="*60)
    print("TEST 1: Detect Complex Fields")
    print("="*60)

    docs = create_test_documents()
    result = ExportService._detect_complex_fields(docs)

    print(f"✓ Detected complex fields: {json.dumps(result, indent=2)}")

    # Assertions
    assert "tags" in result["array"], "Should detect 'tags' as array"
    assert "colors" in result["array"], "Should detect 'colors' as array"
    assert "grading_table" in result["table"], "Should detect 'grading_table' as table"
    assert "line_items" in result["array_of_objects"], "Should detect 'line_items' as array_of_objects"

    print("✓ All complex fields detected correctly\n")
    return True


def test_documents_to_records():
    """Test: Convert documents with complex fields to records"""
    from app.services.export_service import ExportService

    print("="*60)
    print("TEST 2: Documents to Records (Wide Format)")
    print("="*60)

    docs = create_test_documents()
    records = ExportService.documents_to_records(docs)

    print(f"✓ Generated {len(records)} records")

    # Check first record (invoice with complex fields)
    rec1 = records[0]
    print(f"\nRecord 1 fields:")
    print(f"  - invoice_number: {rec1.get('invoice_number')} (type: {type(rec1.get('invoice_number')).__name__})")
    print(f"  - tags: {rec1.get('tags')} (type: {type(rec1.get('tags')).__name__})")
    print(f"  - line_items: {rec1.get('line_items')[:50]}... (type: {type(rec1.get('line_items')).__name__})")

    # Assertions
    assert rec1["invoice_number"] == "INV-001", "Simple field should be string"
    assert isinstance(rec1["tags"], list), "Array field should be list"
    assert len(rec1["tags"]) == 3, "Array should have 3 items"
    assert isinstance(rec1["line_items"], list), "Array of objects should be list"
    assert len(rec1["line_items"]) == 2, "Array of objects should have 2 items"

    print("✓ Records contain correct data types\n")
    return True


def test_create_table_sheet():
    """Test: Create DataFrame for table field"""
    from app.services.export_service import ExportService

    print("="*60)
    print("TEST 3: Create Table Sheet")
    print("="*60)

    docs = create_test_documents()
    df = ExportService._create_table_sheet(docs, "grading_table")

    print(f"✓ Created DataFrame with shape: {df.shape}")
    print(f"\nTable columns: {list(df.columns)}")
    print(f"\nFirst row:\n{df.iloc[0].to_dict()}")

    # Assertions
    assert not df.empty, "DataFrame should not be empty"
    assert "document_id" in df.columns, "Should have document_id column"
    assert "filename" in df.columns, "Should have filename column"
    assert "pom_code" in df.columns, "Should have table column 'pom_code'"
    assert "size_2" in df.columns, "Should have table column 'size_2'"
    assert len(df) == 3, "Should have 3 rows (from table)"
    assert df.iloc[0]["pom_code"] == "B510", "First row should have correct data"

    print("✓ Table sheet created correctly\n")
    return True


def test_create_array_of_objects_sheet():
    """Test: Create DataFrame for array_of_objects field"""
    from app.services.export_service import ExportService

    print("="*60)
    print("TEST 4: Create Array of Objects Sheet")
    print("="*60)

    docs = create_test_documents()
    df = ExportService._create_array_of_objects_sheet(docs, "line_items")

    print(f"✓ Created DataFrame with shape: {df.shape}")
    print(f"\nArray columns: {list(df.columns)}")
    print(f"\nFirst row:\n{df.iloc[0].to_dict()}")

    # Assertions
    assert not df.empty, "DataFrame should not be empty"
    assert "document_id" in df.columns, "Should have document_id column"
    assert "filename" in df.columns, "Should have filename column"
    assert "description" in df.columns, "Should have object property 'description'"
    assert "quantity" in df.columns, "Should have object property 'quantity'"
    assert len(df) == 2, "Should have 2 rows (from array)"
    assert df.iloc[0]["description"] == "Widget A", "First row should have correct data"

    print("✓ Array of objects sheet created correctly\n")
    return True


def test_excel_multi_sheet():
    """Test: Export to Excel with multiple sheets"""
    from app.services.export_service import ExportService

    print("="*60)
    print("TEST 5: Excel Multi-Sheet Export")
    print("="*60)

    docs = create_test_documents()
    records = ExportService.documents_to_records(docs)

    # Export with complex field expansion
    excel_bytes = ExportService.export_to_excel(
        records,
        sheet_name="Main Data",
        documents=docs,
        expand_complex_fields=True
    )

    print(f"✓ Generated Excel file ({len(excel_bytes)} bytes)")

    # Load and inspect the Excel file
    wb = load_workbook(BytesIO(excel_bytes))
    sheet_names = wb.sheetnames

    print(f"\nSheet names: {sheet_names}")

    # Assertions
    assert "Main Data" in sheet_names, "Should have main data sheet"
    assert any("grading_table" in name for name in sheet_names), "Should have grading_table sheet"
    assert any("line_items" in name for name in sheet_names), "Should have line_items sheet"

    # Check main sheet
    main_sheet = wb["Main Data"]
    headers = [cell.value for cell in main_sheet[1]]
    print(f"\nMain sheet headers (first 10): {headers[:10]}")

    assert "document_id" in headers, "Main sheet should have document_id"
    assert "filename" in headers, "Main sheet should have filename"
    assert "invoice_number" in headers or any("invoice" in str(h).lower() for h in headers), "Main sheet should have field columns"

    # Check table sheet
    table_sheet_name = [name for name in sheet_names if "grading_table" in name][0]
    table_sheet = wb[table_sheet_name]
    table_headers = [cell.value for cell in table_sheet[1]]
    print(f"\nTable sheet '{table_sheet_name}' headers: {table_headers}")

    assert "pom_code" in table_headers, "Table sheet should have table columns"

    print("✓ Excel file has correct structure with multiple sheets\n")
    return True


def test_csv_export():
    """Test: CSV export with complex fields as strings"""
    from app.services.export_service import ExportService

    print("="*60)
    print("TEST 6: CSV Export (Complex fields as JSON strings)")
    print("="*60)

    docs = create_test_documents()
    records = ExportService.documents_to_records(docs)

    # Export to CSV
    csv_bytes = ExportService.export_to_csv(records, include_metadata=True)
    csv_text = csv_bytes.decode('utf-8')

    print(f"✓ Generated CSV ({len(csv_bytes)} bytes)")
    print(f"\nFirst 500 characters:\n{csv_text[:500]}")

    # Parse CSV to verify
    df = pd.read_csv(BytesIO(csv_bytes))

    # Assertions
    assert not df.empty, "CSV should not be empty"
    assert "document_id" in df.columns, "CSV should have document_id"
    assert "filename" in df.columns, "CSV should have filename"

    # Check that complex fields are present (as JSON strings or serialized)
    row1 = df.iloc[0]
    if "tags" in df.columns:
        tags_value = row1["tags"]
        print(f"\n✓ Tags field in CSV: {tags_value} (type: {type(tags_value).__name__})")
        # Could be list or JSON string depending on serialization

    print("✓ CSV export successful\n")
    return True


def test_json_export():
    """Test: JSON export preserves complex structures"""
    from app.services.export_service import ExportService

    print("="*60)
    print("TEST 7: JSON Export (Native structures)")
    print("="*60)

    docs = create_test_documents()
    records = ExportService.documents_to_records(docs)

    # Export to JSON
    json_str = ExportService.export_to_json(records, format_type="pretty")
    json_data = json.loads(json_str)

    print(f"✓ Generated JSON ({len(json_str)} bytes)")
    print(f"\nFirst record (truncated):\n{json.dumps(json_data[0], indent=2)[:500]}...")

    # Assertions
    assert isinstance(json_data, list), "JSON should be array"
    assert len(json_data) == 3, "Should have 3 records"

    rec1 = json_data[0]
    assert rec1["invoice_number"] == "INV-001", "Simple field should be preserved"
    assert isinstance(rec1["tags"], list), "Array should be native list in JSON"
    assert isinstance(rec1["line_items"], list), "Array of objects should be native list"
    assert isinstance(rec1["line_items"][0], dict), "Objects should be native dicts"

    print("✓ JSON export preserves native structures\n")
    return True


def run_all_tests():
    """Run all validation tests"""
    print("\n" + "█"*60)
    print("EXPORT REFACTOR VALIDATION TESTS")
    print("█"*60)

    tests = [
        ("Detect Complex Fields", test_detect_complex_fields),
        ("Documents to Records", test_documents_to_records),
        ("Create Table Sheet", test_create_table_sheet),
        ("Create Array of Objects Sheet", test_create_array_of_objects_sheet),
        ("Excel Multi-Sheet Export", test_excel_multi_sheet),
        ("CSV Export", test_csv_export),
        ("JSON Export", test_json_export),
    ]

    results = []
    for name, test_func in tests:
        try:
            result = test_func()
            results.append((name, "PASS" if result else "FAIL"))
        except Exception as e:
            print(f"\n❌ ERROR in {name}: {str(e)}")
            import traceback
            traceback.print_exc()
            results.append((name, "ERROR"))

    # Summary
    print("\n" + "="*60)
    print("TEST SUMMARY")
    print("="*60)

    for name, status in results:
        icon = "✓" if status == "PASS" else "❌"
        print(f"{icon} {name}: {status}")

    passed = sum(1 for _, status in results if status == "PASS")
    total = len(results)

    print(f"\n{passed}/{total} tests passed")

    if passed == total:
        print("\n🎉 ALL TESTS PASSED! Backend export refactor is ready for frontend integration.")
        return 0
    else:
        print("\n⚠️  Some tests failed. Please review the errors above.")
        return 1


if __name__ == "__main__":
    sys.exit(run_all_tests())
